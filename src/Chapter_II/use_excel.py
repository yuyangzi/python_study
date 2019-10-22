from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


class ExcelUtils(object):

    def __init__(self):
        self.work_book = Workbook()
        self.work_sheet0 = self.work_book.active
        self.work_sheet1 = self.work_book.create_chartsheet(title='诗人')

    def do_sth(self):
        self.work_sheet0.title = 'repast'
        self.work_sheet0.sheet_properties.tabColor = 'FF0000'
        self.work_sheet0['A1'] = 'pea 豆子'
        self.work_sheet0['A2'] = 'bean 豆类'

        for row in self.work_sheet0['B1:F7']:
            for cell in row:
                cell.value = datetime.now().date()

        self.work_sheet0['G1'] = '=SUM(B1:F1)'

        image = Image('./static/5b3307f50001859e05400300-280-160.jpg')
        self.work_sheet0.add_image(image, 'G19')

        self.work_book.save('./static/test.xlsx')

    @staticmethod
    def read_xls():
        xls = load_workbook('./static/template.xlsx')
        names = xls.sheetnames
        work = xls[names[0]]
        print(work.title)
        print(names)

        for row in work['A1:C12']:
            for cell in row:
                print(cell.value)


if __name__ == '__main__':
    excel_util = ExcelUtils()
    excel_util.read_xls()
